"""Export the Excel asset ledger to ``CMDB/CMDB.csv``.

Reads the physical Excel ledger (default sheet ``管理番号ﾍﾞｰｽ``), matches
columns by normalized header name (not fixed index), applies the deploy-year
and asset-number filters, and writes a UTF-8-BOM CSV suitable for downstream
import. Any existing CSV is backed up first.

Can be run as a module::

    python -m cmdb.export_cmdb_csv --input "../CMDB/■情報機器管理台帳.xlsx"
"""

import argparse
import csv
import datetime
import os
import shutil
import warnings

try:  # package-relative import (normal case)
    from .header_map import CSV_COLUMNS, LEDGER_HEADERS, normalize
except ImportError:  # pragma: no cover - fallback for direct script execution
    from header_map import CSV_COLUMNS, LEDGER_HEADERS, normalize

__all__ = ["export_cmdb_csv"]

DEFAULT_INPUT = "CMDB/■情報機器管理台帳.xlsx"
DEFAULT_OUTPUT = "CMDB/CMDB.csv"
DEFAULT_SHEET = "管理番号ﾍﾞｰｽ"
DEFAULT_BACKUP_DIR = "CMDB/Data-BackUp"
DEFAULT_MIN_YEAR = 2019


def _parse_year(value):
    """Return the leading 4-digit year as int, or ``None`` if absent.

    Accepts ``2019``, ``"2019"``, ``"2019年度"`` etc. Returns ``None`` when the
    value is empty / non-numeric so callers can treat "year present" vs
    "year missing" distinctly.
    """
    if value is None:
        return None
    text = normalize(value)
    if not text:
        return None
    digits = ""
    for ch in text:
        if ch.isdigit():
            digits += ch
        else:
            break
    if len(digits) >= 4:
        try:
            return int(digits[:4])
        except ValueError:
            return None
    return None


def _build_header_index(header_row):
    """Map logical keys -> column index from a raw header row.

    Matches by normalized header name so embedded newlines / width
    differences do not break detection.
    """
    index = {}
    for col_idx, cell in enumerate(header_row):
        key = LEDGER_HEADERS.get(normalize(cell))
        if key is not None and key not in index:
            index[key] = col_idx
    return index


def _cell(row, idx):
    """Safely fetch ``row[idx]`` returning ``None`` when out of range."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _stringify(value):
    """Return a trimmed string for CSV output, "" for ``None``."""
    if value is None:
        return ""
    return str(value).strip()


def export_cmdb_csv(
    xlsx_path,
    csv_path=DEFAULT_OUTPUT,
    sheet_name=DEFAULT_SHEET,
    min_year=DEFAULT_MIN_YEAR,
    backup_dir=DEFAULT_BACKUP_DIR,
):
    """Export the ledger Excel file to a CMDB CSV.

    Args:
        xlsx_path: Path to the source ``.xlsx`` ledger.
        csv_path: Output CSV path (overwritten; backed up first if present).
        sheet_name: Worksheet to read. If absent, an error listing available
            sheets is raised.
        min_year: Rows whose deploy year is below this are skipped. Rows with
            no year but a non-empty asset number are still emitted.
        backup_dir: Directory to copy an existing CSV into before overwrite.

    Returns:
        dict with ``rows`` (written), ``skipped``, and ``backup`` (path or
        ``None``).
    """
    import openpyxl  # imported lazily so the package loads without openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"sheet {sheet_name!r} not found in {xlsx_path!r}; "
                f"available sheets: {available}"
            )
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            header_row = ()
        hidx = _build_header_index(header_row)

        out_rows = []
        skipped = 0
        for raw in rows_iter:
            if raw is None:
                skipped += 1
                continue
            asset_number = _stringify(_cell(raw, hidx.get("asset_number")))
            if not asset_number:
                # No 管理番号 -> not an asset row.
                skipped += 1
                continue
            year = _parse_year(_cell(raw, hidx.get("deploy_year")))
            if year is not None and year < min_year:
                skipped += 1
                continue
            out_rows.append(
                [
                    "" if year is None else str(year),
                    asset_number,
                    "",  # IP (LAN/WiFi) — back-filled in Phase I-2
                    _stringify(_cell(raw, hidx.get("owner_name"))),
                    _stringify(_cell(raw, hidx.get("employee_id"))),
                    _stringify(_cell(raw, hidx.get("os"))),
                    _stringify(_cell(raw, hidx.get("mac_wired"))),
                    _stringify(_cell(raw, hidx.get("mac_wireless"))),
                ]
            )
    finally:
        wb.close()

    backup_path = None
    if os.path.exists(csv_path):
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        backup_path = os.path.join(backup_dir, f"CMDB_{ts}.csv")
        shutil.copy2(csv_path, backup_path)

    out_dir = os.path.dirname(csv_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(CSV_COLUMNS)
        writer.writerows(out_rows)

    try:
        os.chmod(csv_path, 0o600)
    except OSError as exc:  # pragma: no cover - platform dependent
        warnings.warn(f"could not chmod {csv_path}: {exc}", stacklevel=2)

    return {"rows": len(out_rows), "skipped": skipped, "backup": backup_path}


def _main(argv=None):
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Export Excel ledger to CMDB CSV")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="source .xlsx path")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="output .csv path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="worksheet name")
    parser.add_argument(
        "--min-year", type=int, default=DEFAULT_MIN_YEAR, help="minimum deploy year"
    )
    args = parser.parse_args(argv)
    result = export_cmdb_csv(
        args.input,
        csv_path=args.output,
        sheet_name=args.sheet,
        min_year=args.min_year,
    )
    print(
        f"wrote {result['rows']} rows to {args.output} "
        f"(skipped {result['skipped']}, backup {result['backup']})"
    )
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(_main())
